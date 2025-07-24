
import logging
import os
from typing import Any, Iterable, AsyncIterable, Sequence, Union

import xlsxwriter
from openpyxl.styles import PatternFill, Font


class StreamingExcelExporter:
    def __init__(
        self,
        header_format_opts = {"bold": True, "bg_color": "#DDDDDD"},
        header_fill_color: str = "DDDDDD",
        header_font_bold: bool = True,
        min_col_width: int = 10,
        max_col_width: int = 50

    ):
        self.header_fill = PatternFill(
            start_color=header_fill_color,
            end_color=header_fill_color,
            fill_type="solid"
        )
        self.header_format_opts = header_format_opts
        self.header_font = Font(bold=header_font_bold)
        self.min_col_width = min_col_width
        self.max_col_width = max_col_width

    async def export(
        self,
        data: Union[Iterable[Sequence[Any]], AsyncIterable[Sequence[Any]]],
        headers: Sequence[str],
        sheet_name: str,
        file_path: str
    ):
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        workbook = xlsxwriter.Workbook(file_path, {"constant_memory": True})
        worksheet = workbook.add_worksheet(sheet_name or "Sheet1")
        header_fmt = workbook.add_format(self.header_format_opts)
        max_widths = []
        try:

            for col_idx, h in enumerate(headers):
                worksheet.write(0, col_idx, h, header_fmt)
                max_widths.append(len(str(h)))

            row_idx = 1

            if hasattr(data, "__aiter__"):
                async for row in data:
                    for col_idx, value in enumerate(row):
                        worksheet.write(row_idx, col_idx, value)
                        length = len(str(value or ""))
                        if length > max_widths[col_idx]:
                            max_widths[col_idx] = length
                    row_idx += 1
            else:
                for row in data:
                    for col_idx, value in enumerate(row):
                        worksheet.write(row_idx, col_idx, value)
                        length = len(str(value or ""))
                        if length > max_widths[col_idx]:
                            max_widths[col_idx] = length
                    row_idx += 1

        except Exception as e:
            logging.error(e)
            raise e

        finally:
            for col_idx, w in enumerate(max_widths):
                width = min(max(w + 2, self.min_col_width), self.max_col_width)
                worksheet.set_column(col_idx, col_idx, width)
            workbook.close()
