import io
from openpyxl import Workbook
from typing import Any, List, Dict, Union, Tuple, Sequence, Mapping

from openpyxl.styles import PatternFill

from file_export.export import Export


class ExportExcel(Export):

    def autofit_columns(self, ws, min_width: int = 10, max_width: int = 50):

        for col in ws.columns:
            col_letter = col[0].column_letter
            lengths = []
            for cell in col:
                if cell.value is not None:
                    lengths.append(len(str(cell.value)))
            if lengths:
                best = max(lengths)
                width = max(min_width, min(best + 2, max_width))
                ws.column_dimensions[col_letter].width = width

    def apply_header_fill(self, ws, header_row: int = 1, fill_color: str = "DDDDDD"):

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for cell in ws[header_row]:
            cell.fill = fill

    def export(self, data: Union[Dict[Any, Union[Dict, List, Tuple]], Sequence[Sequence[Any]], Tuple], **options):
        if isinstance(data, Mapping):
            datasets = [data]
        elif isinstance(data, Sequence) and not isinstance(data, (str, bytes, Mapping)) and data and all(
                isinstance(r, Sequence) and not isinstance(r, (str, bytes, Mapping)) for r in data):
            datasets = [data]
        elif isinstance(data, Sequence) and not isinstance(data, (str, bytes, Mapping)):
            datasets = list(data)
        else:
            raise TypeError("`data` must be a dict, a sequence of rows, or a tuple/list of those")

        wb = Workbook()
        raw_names = options.get("sheet_name") or []
        sheet_names = raw_names if isinstance(raw_names, list) else [raw_names]
        raw_h = options.get("headers") or []
        headers = raw_h if all(isinstance(el, (list, tuple)) for el in raw_h) else [raw_h]
        raw_k = options.get("key_header")
        key_headers = raw_k if isinstance(raw_k, list) else [raw_k] if raw_k is not None else []
        file_path = options["file_path"]

        for idx, dataset in enumerate(datasets):
            ws = wb.active if idx == 0 else wb.create_sheet()
            if idx < len(sheet_names) and sheet_names[idx]:
                ws.title = sheet_names[idx]

            hdrs = headers[idx] if idx < len(headers) else None
            kh = key_headers[idx] if idx < len(key_headers) else None

            if isinstance(dataset, Mapping):
                vals = list(dataset.values())
                first = vals[0] if vals else None
                if first is None:
                    raise ValueError(f"Sheet #{idx} dataset is empty")
                if isinstance(first, Mapping):
                    cols = list(first.keys())
                else:
                    cols = [f"Col{i + 1}" for i in range(len(first))]
                if hdrs:
                    if len(hdrs) != len(cols):
                        raise ValueError("len(headers) != number of columns")
                    cols = hdrs
                ws.append([kh, *cols] if kh else cols)

                for key, rowval in dataset.items():
                    if isinstance(rowval, Mapping):
                        row = [rowval.get(c) for c in cols]
                    else:
                        row = list(rowval)
                    if kh:
                        row = [key, *row]
                    ws.append(row)
            else:
                rows = list(dataset)
                ncols = max(len(r) for r in rows)
                cols = [f"Col{i + 1}" for i in range(ncols)]
                if hdrs:
                    if len(hdrs) != ncols:
                        raise ValueError("len(headers) != number of columns")
                    cols = hdrs
                ws.append(cols)
                for row in rows:
                    r = list(row)
                    if len(r) < ncols:
                        r += [None] * (ncols - len(r))
                    ws.append(r)

            self.autofit_columns(ws)
            self.apply_header_fill(ws)

        wb.save(file_path)
