import aiofiles
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from typing import AsyncIterable, Sequence


class ExportExcelAsync:
    async def export(
        self,
        rows_async_iter: AsyncIterable[Sequence],
        headers: Sequence[str],
        sheet_name: str,
        file_path: str
    ) -> None:

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name)

        ws.append(list(headers))
        fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws.rows.__next__():
            cell.fill = fill

        async for row in rows_async_iter:
            ws.append(list(row))

        virtual_wb = wb.save
        from openpyxl.writer.excel import save_virtual_workbook
        data = save_virtual_workbook(wb)

        async with aiofiles.open(file_path, "wb") as f:
            await f.write(data)
